import openpyxl
from openpyxl import Workbook
wb = Workbook()

wb = openpyxl.load_workbook('srnames.xlsx')
if 'surn_sort' not in wb.sheetnames:
    wb.create_sheet('surn_sort')
wb.save('srnames.xlsx')
ws = wb['surn_sort']

with open('surnames.txt', 'r', encoding="utf-8") as sn:
    snr = [snr.rstrip() for snr in sn.readlines()]
    print(snr)
    for i in snr:
        snr = i.split(', ')
        ws.append(snr)
        wb.save('srnames.xlsx')
    print(snr)

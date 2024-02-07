import openpyxl
from openpyxl import Workbook
import statistics
wb = Workbook()

wb = openpyxl.load_workbook('tt.xlsx')
ws = wb.active
dct = {}
res = []

for i in range(1, ws.max_row+1):
    res.append(ws.cell(row=i, column=2).value)

dct['Максимальное значение'] = max(res)
dct['Минимальное значение'] = min(res)
dct['Summa'] = sum(res)
dct['Srednee'] = statistics.mean(res)
dct['Mediana'] = statistics.median(res)
print(dct)

ws = wb.create_sheet('Statistics')

i = 1
for j in dct.keys():
    ws.cell(row=i, column=1).value = j
    i += 1
i = 1
for k in dct:
    ws.cell(row=i, column=2).value = dct[k]
    i += 1
wb.save('tt.xlsx')
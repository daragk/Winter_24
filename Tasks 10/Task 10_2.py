import openpyxl
from openpyxl import Workbook

res = []
dct = {}
wb = Workbook()
wb = openpyxl.load_workbook('test.xlsx')
wb.create_sheet('Sorted')
#wb.save('test.xlsx')

ws = wb.active
for i in range(1,ws.max_row+1):
    dct[ws.cell(row=i, column=1).value] = ws.cell(row=i, column=2).value
sm = sum(dct.values())
sdct = sorted(dct.items(), key= lambda x: (x[1], x[0]), reverse=True)

ws = wb['Sorted']
for i in range(1, len(sdct) + 1):
    ws.cell(row=i, column=1).value = sdct[i - 1][0]
    ws.cell(row=i, column=2).value = sdct[i - 1][1]


ws.cell(row=ws.max_row+1, column=1).value = 'Total'
ws.cell(row=ws.max_row, column=2).value = sm

wb.save('test.xlsx')
